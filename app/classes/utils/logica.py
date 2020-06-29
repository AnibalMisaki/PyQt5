from classes.utils.logger import logger
from classes.objects.workSpace import workSpace,device, variable
from ..objects.usuario import usuario
from ..objects.reporte import reporte
from py_expression_eval import Parser
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Border, Side
import requests, json, os

class Logica():

    __program_files = '/opt' #os.environ["ALLUSERSPROFILE"]

    settings = { 
            "Host":"127.0.0.1",
            "Port":8080,
            "HTTP_PROTOCOL":"http",
            "URI":"http://127.0.0.1:8080"
    }
    
    @staticmethod
    def IniciarSesion(**kwargs):
        _data = {"Usuario":kwargs["Usuario"],"Password":kwargs["Password"]}
        _response =  requests.post("%s/Sesion/IniciarSesion" % (Logica.settings["URI"]), timeout = 45,json=_data)
        return json.loads(_response.content,object_hook=usuario)

    @staticmethod
    def ObtenerProyectos(**kwargs): # returns a list with all project in database
        _headers = {'Authorization': 'Bearer ' + kwargs["access_token"]}
        result  = requests.get("%s/Controles/MostrarTodos" % (Logica.settings["URI"]), timeout = 45, headers=_headers)
        result.raise_for_status()
        _answer:workSpace = json.loads( result.content ,object_hook=workSpace)
        return _answer

    @staticmethod
    def ObtenerWorkspaces(**kwargs):
        _headers = {'Authorization': 'Bearer ' + kwargs["access_token"]}
        response  = requests.get("%s/Controles/ObtenerTodos" % (Logica.settings["URI"]), timeout = 45, headers=_headers)
        response.raise_for_status()
        return list(Logica.jsonToList(response.json(),workSpace))

    @staticmethod
    def ObtenerConfiguraciones(**kwargs):
        _headers = {'Authorization': 'Bearer ' + kwargs["access_token"]}
        result  = requests.get("%s/Configuraciones/Obtener" % (Logica.settings["URI"]), timeout = 45, headers=_headers)
        result.raise_for_status()
        return result.json()

    @staticmethod
    def GuardarConfiguraciones(**kwargs):
        _headers = {'Authorization': 'Bearer ' + kwargs["access_token"]}
        result  = requests.post("%s/Configuraciones/Actualizar" % (Logica.settings["URI"]), timeout = 45,json=kwargs["data"], headers=_headers)
        result.raise_for_status()
        return result.json()
        
    @staticmethod
    def AbrirProyecto(**kwargs):
        _headers = {'Authorization': 'Bearer ' + kwargs["access_token"]}
        response  = requests.get("%s/Controles/Abrir/%s" % (Logica.settings["URI"],kwargs["id"]), timeout = 45, headers=_headers)
        response.raise_for_status()
        if response.json() == []:
            raise Exception("¡Error! Proyecto no existe")
        return workSpace(response.json())

    @staticmethod
    def ObtenerVariablesFunciones(**kwargs):
        _headers = {'Authorization': 'Bearer ' + kwargs["access_token"]}
        result  = requests.get("%s/Controles/ObtenerVariablesFunciones/%s/%s" % (Logica.settings["URI"],kwargs["ID"],kwargs["Token"]), timeout = 45, headers=_headers)
        result.raise_for_status()
        return result.json()

    @staticmethod
    def LeerSensor(**kwargs):
        _headers = {'Authorization': 'Bearer ' + kwargs["access_token"]}
        result  = requests.post("%s/Controles/LeerSensor/%s/%s" % (Logica.settings["URI"],kwargs["ID"],kwargs["Token"]), timeout = 45,json=kwargs["data"],headers=_headers)
        result.raise_for_status()
        _answer:variable = json.loads(result.content, object_hook=variable)
        return _answer

    @staticmethod
    def ActualizarSensor(**kwargs):
        _headers = {'Authorization': 'Bearer ' + kwargs["access_token"]}
        result  = requests.post("%s/Controles/ActualizarVariable/%s/%s" % (Logica.settings["URI"],kwargs["ID"],kwargs["Token"]), timeout = 45,json=kwargs["data"],headers=_headers)
        result.raise_for_status()
        return result.json()

    @staticmethod
    def Guardar(**kwargs):
        _headers = {'Authorization': 'Bearer ' + kwargs["access_token"]}
        result  = requests.post("%s/Controles/Guardar" % (Logica.settings["URI"]), timeout = 45,json=kwargs["data"],headers=_headers)
        result.raise_for_status()
        return result.json()

    @staticmethod
    def EliminarProyecto(**kwargs):
        _headers = {'Authorization': 'Bearer ' + kwargs["access_token"]}
        result  = requests.get("%s/Controles/Eliminar/%s" % (Logica.settings["URI"],kwargs["id"]), timeout = 45, headers=_headers)
        result.raise_for_status()
        return result.json()

    @staticmethod
    def ObtenerUsuarios(**kwargs):
        _headers = {'Authorization': 'Bearer ' + kwargs["access_token"]}
        response  = requests.get("%s/Sesion/ObtenerUsuarios" % (Logica.settings["URI"]), timeout = 45, headers=_headers)
        response.raise_for_status()
        return list(Logica.jsonToList(response.json(),usuario))

    @staticmethod
    def agregarUsuario(**kwargs):
        _headers = {'Authorization': 'Bearer ' + kwargs["access_token"]}
        response  = requests.post("%s/Sesion/InsertarUsuario" % (Logica.settings["URI"]), timeout = 45,json=kwargs["data"], headers=_headers)
        response.raise_for_status()
        return response.json()

    @staticmethod
    def editarUsuario(**kwargs):
        _headers = {'Authorization': 'Bearer ' + kwargs["access_token"]}
        response  = requests.post("%s/Sesion/ActualizarUsuario" % (Logica.settings["URI"]), timeout = 45,json=kwargs["data"], headers=_headers)
        response.raise_for_status()
        return response.json()
    
    @staticmethod
    def actualizarUsuario(**kwargs):
        _headers = {'Authorization': 'Bearer ' + kwargs["access_token"]}
        response  = requests.post("%s/Sesion/ActualizarCuenta" % (Logica.settings["URI"]), timeout = 45,json=kwargs["data"], headers=_headers)
        response.raise_for_status()
        return response.json()

    @staticmethod
    def actualizarContrasenia(**kwargs):
        _headers = {'Authorization': 'Bearer ' + kwargs["access_token"]}
        response  = requests.post("%s/Sesion/CambiarContrasenia" % (Logica.settings["URI"]), timeout = 45,json=kwargs["data"], headers=_headers)
        response.raise_for_status()
        return response.json()

    @staticmethod
    def recuperarContrasenia(**kwargs):
        response  = requests.post("%s/Sesion/RecuperarContrasenia" % (Logica.settings["URI"]), timeout = 45,json=kwargs["data"])
        response.raise_for_status()
        return response.json()

    @staticmethod
    def eliminarUsuario(**kwargs):
        _headers = {'Authorization': 'Bearer ' + kwargs["access_token"]}
        response  = requests.get("%s/Sesion/EliminarUsuario/%s" % (Logica.settings["URI"],kwargs["id"]), timeout = 45, headers=_headers)
        response.raise_for_status()
        return response.json()

    @staticmethod
    def habilitarUsuario(**kwargs):
        _headers = {'Authorization': 'Bearer ' + kwargs["access_token"]}
        response  = requests.get("%s/Sesion/HabilitarUsuario/%s" % (Logica.settings["URI"],kwargs["id"]), timeout = 45, headers=_headers)
        response.raise_for_status()
        return response.json()
    
    @staticmethod
    def deshabilitarUsuario(**kwargs):
        _headers = {'Authorization': 'Bearer ' + kwargs["access_token"]}
        response  = requests.get("%s/Sesion/DeshabilitarUsuario/%s" % (Logica.settings["URI"],kwargs["id"]), timeout = 45, headers=_headers)
        response.raise_for_status()
        return response.json()

    @staticmethod
    def nuevoReporte(**kwargs):
        _headers = {'Authorization': 'Bearer ' + kwargs["access_token"]}
        result  = requests.post("%s/Reportes/NuevoReporte" % (Logica.settings["URI"]), timeout = 45,json=kwargs["data"], headers=_headers)
        result.raise_for_status()
        return result.json()

    @staticmethod
    def ObtenerReportes(**kwargs):
        _headers = {'Authorization': 'Bearer ' + kwargs["access_token"]}
        response  = requests.get("%s/Reportes/ObtenerReportes/%s/%s/%s" % (Logica.settings["URI"],kwargs["dateStart"],kwargs["dateEnd"],kwargs["nivel"]), timeout = 45, headers=_headers)
        response.raise_for_status()
        return list(Logica.jsonToList(response.json(),reporte))

    @staticmethod
    def LeerConfiguracionesLocal():
        try:
            with open("%s/Sistema SCADA/setting.json" % Logica.__program_files,'r') as f:
                jsonFile = json.load(f)
                f.close()
                return jsonFile
        except Exception as e:
            return None
        
    @staticmethod
    def GuardarConfiguracionesLocal(settings):
        try:
            with open("%s/Sistema SCADA/setting.json" % Logica.__program_files,"w") as f:
                f.write(json.dumps(settings))
                f.close()
            return True
        except Exception as e:
            Logger = logger()
            Logger.log_error(e)
            return False

    @staticmethod
    def ExportarReportes(filename,reportes:list):
        Logger = logger()
        workbook = Workbook()
        currentShet = workbook.active
        currentShet.title = "Pagina 1"
        for i,reporte in enumerate(reportes):
            border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
            filler = PatternFill(fill_type='solid', start_color=Logica.findHEX(reporte.nivel),end_color=Logica.findHEX(reporte.nivel))
            currentShet["A%s" % (i+1).__str__()].fill = filler
            currentShet["A%s" % (i+1).__str__()].border = border
            currentShet["B%s" % (i+1).__str__()] = reporte.nombreVariable
            currentShet["C%s" % (i+1).__str__()] = reporte.nombreDispositivo
            currentShet["D%s" % (i+1).__str__()] = reporte.fecha
            currentShet["E%s" % (i+1).__str__()] = reporte.hora
            currentShet["F%s" % (i+1).__str__()] = reporte.condicion
            currentShet["G%s" % (i+1).__str__()] = reporte.usuario
            currentShet["H%s" % (i+1).__str__()] = reporte.value.__str__()
            currentShet["I%s" % (i+1).__str__()] = reporte.mensaje
        try:
            import os
            workbook.save(filename=filename)
            fileNameStr = os.path.splitext(filename)[0]
            os.rename(filename, fileNameStr + '.xlsx') 
            return True
        except Exception as e:
            Logger.log_error(e)
            return False


    @staticmethod
    def evaluarExpresion(expresion:str,value=255):
        parser = Parser()
        try:
            response = parser.parse(expresion).evaluate({'x':value})
            return response
        except Exception as e:
            return False

    @staticmethod
    def findHEX(color:str):
        switcher = {
            "aqua":"00FFFF",
            "green":"008000",
            "orange":"FFA500",
            "red":"FF0000"
        }
        return switcher.get(color,"aqua")
    
    @staticmethod
    def jsonToList(y:list,type):
        if y is None or len(y) == 0:
            return []
        for x in y:
            yield type(x)

    @staticmethod
    def fileToJSON(fileName):
        Logger = logger()
        try:
            with open(fileName,'r') as f:
                file = json.loads(f.read())
            return file
        except Exception as e:
            Logger.log_error(e)
            return None

    @staticmethod
    def JSONToFile(fileName:str,jsonData):
        Logger = logger()
        try:
            import os
            with open(fileName,'w+') as f:
                f.write(json.dumps(jsonData))
                f.close()
            fileNameStr = os.path.splitext(fileName)[0]
            os.rename(fileName, fileNameStr + '.json')
            return True
        except Exception as e:
            Logger.log_error(e)
            return False

    @staticmethod
    def imageToByteArray(fileName):
        Logger = logger()
        from base64 import b64encode
        try:
            with open(fileName,'rb') as f:
                file = b64encode(f.read())
            return file.decode()
        except Exception as e:
            Logger.log_error(e)
            return None
        
    @staticmethod
    def byteArrayToImage(file):
        from base64 import b64encode, b64decode
        from PyQt5.QtGui import QPixmap, QImage
        from PyQt5.QtCore import QByteArray
        return QPixmap.fromImage(QImage.fromData(QByteArray.fromBase64 ( bytes(file,"utf-8") )))

    @staticmethod
    def LeerConfiguracion():
        Logger = logger()
        try:
            with open("%s/Sistema SCADA/setting.json" % Logica.__program_files,"r") as f:
                jsonFile = json.load(f)
                jsonFile["URI"] = "%s://%s" % (jsonFile["HTTP_PROTOCOL"],jsonFile["Host"]) if jsonFile["Port"] is None else "%s://%s:%s" % (jsonFile["HTTP_PROTOCOL"],jsonFile["Host"],jsonFile["Port"])
                f.close()
                return jsonFile
        except Exception as e: # Si no existe lo crea
            Logger.log_error( Exception("¡Error! No pudo leerse el archivo de configuracion, Generando uno nuevo"))
            settings =  {
                "Host":"127.0.0.1",
                "Port":8080,
                "HTTP_PROTOCOL":"http"
            }
            with open("%s/Sistema SCADA/setting.json" % Logica.__program_files,"w") as f:
                f.write(json.dumps(settings))
                f.close()
            settings["URI"] = "%s://%s" % (settings["HTTP_PROTOCOL"],settings["Host"]) if settings["Port"] is None else "%s://%s:%s" % (settings["HTTP_PROTOCOL"],settings["Host"],settings["Port"])
            return settings

Logica.settings = Logica.LeerConfiguracion()

    

